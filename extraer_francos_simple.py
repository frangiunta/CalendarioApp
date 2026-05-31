#!/usr/bin/env python3
"""
Script simplificado para extraer francos de GIGANET y generar Excel compatible.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re

ARCHIVO_ORIGINAL = '/workspaces/CalendarioApp/app-francos/src/app/3.Nueva Grilla 2026 GIGANET.xlsx'
ARCHIVO_SALIDA = '/workspaces/CalendarioApp/app-francos/src/assets/francos.xlsx'

def main():
    print("="*80)
    print("🔄 EXTRAYENDO FRANCOS DEL EXCEL")
    print("="*80)
    
    wb_original = openpyxl.load_workbook(ARCHIVO_ORIGINAL)
    
    # Hojas con fechas
    hojas_relevantes = [h for h in wb_original.sheetnames if ' AL ' in h]
    print(f"\n✓ Hojas encontradas: {hojas_relevantes}\n")
    
    todos_empleados_francos = {}
    
    for nombre_hoja in hojas_relevantes:
        print(f"📋 Procesando: {nombre_hoja}")
        ws = wb_original[nombre_hoja]
        
        # Extraer fechas de fila 3
        fechas_columnas = {}
        for col_idx in range(3, ws.max_column + 1):
            val = ws.cell(3, col_idx).value
            if val:
                texto = str(val).strip()
                match = re.search(r'(\d{4})-(\d{2})-(\d{2})', texto)
                if match:
                    fecha_str = match.group(0)
                    if fecha_str not in fechas_columnas:
                        fechas_columnas[fecha_str] = []
                    fechas_columnas[fecha_str].append(col_idx)
        
        print(f"  ✓ Fechas: {len(fechas_columnas)}")
        
        # Procesar empleados
        for row_idx in range(7, ws.max_row + 1):
            nombre = ws.cell(row_idx, 2).value
            
            if not nombre or not str(nombre).strip():
                continue
            
            nombre = str(nombre).strip()
            
            # Limpiar nombre
            if nombre[0].isdigit():
                for i, char in enumerate(nombre):
                    if not char.isdigit() and char != ' ':
                        nombre = nombre[i:].strip()
                        break
            
            if not nombre or nombre.upper() in ['*', '']:
                continue
            
            # Buscar francos
            francos = []
            for fecha_key in sorted(fechas_columnas.keys()):
                columnas = fechas_columnas[fecha_key]
                tiene_presente = False
                
                for col_idx in columnas:
                    val = ws.cell(row_idx, col_idx).value
                    if val and str(val).strip().upper() == 'P':
                        tiene_presente = True
                        break
                
                if not tiene_presente:
                    francos.append(fecha_key)
            
            if francos:
                if nombre not in todos_empleados_francos:
                    todos_empleados_francos[nombre] = []
                todos_empleados_francos[nombre].extend(francos)
        
        print(f"  ✓ Empleados con francos: {sum(1 for e in todos_empleados_francos.values() if e)}\n")
    
    # Crear Excel compatible
    print("\n📝 Creando Excel compatible...")
    
    wb_nuevo = openpyxl.Workbook()
    ws_nuevo = wb_nuevo.active
    ws_nuevo.title = "Francos"
    
    # Fechas únicas
    todas_fechas = set()
    for francos in todos_empleados_francos.values():
        todas_fechas.update(francos)
    
    fechas_ordenadas = sorted(list(todas_fechas))
    
    # Encabezados
    ws_nuevo.cell(1, 1, "Nombre")
    for col_idx, fecha in enumerate(fechas_ordenadas, start=2):
        ws_nuevo.cell(1, col_idx, fecha)
    
    # Datos
    for row_idx, (nombre, francos) in enumerate(sorted(todos_empleados_francos.items()), start=2):
        ws_nuevo.cell(row_idx, 1, nombre)
        
        francos_set = set(francos)
        for col_idx, fecha in enumerate(fechas_ordenadas, start=2):
            if fecha in francos_set:
                ws_nuevo.cell(row_idx, col_idx, "F")
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    franco_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    for col_idx in range(1, len(fechas_ordenadas) + 2):
        cell = ws_nuevo.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx in range(2, len(todos_empleados_francos) + 2):
        for col_idx in range(2, len(fechas_ordenadas) + 2):
            cell = ws_nuevo.cell(row_idx, col_idx)
            if cell.value == "F":
                cell.fill = franco_fill
                cell.alignment = Alignment(horizontal="center")
    
    ws_nuevo.column_dimensions['A'].width = 25
    for col_idx in range(2, len(fechas_ordenadas) + 2):
        ws_nuevo.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
    
    wb_nuevo.save(ARCHIVO_SALIDA)
    print(f"✅ Excel generado: {ARCHIVO_SALIDA}")
    print(f"✅ Empleados: {len(todos_empleados_francos)}")
    print(f"✅ Fechas: {len(fechas_ordenadas)}")
    print(f"✅ Total de francos: {sum(len(f) for f in todos_empleados_francos.values())}")

if __name__ == "__main__":
    main()
