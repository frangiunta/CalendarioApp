#!/usr/bin/env python3
"""
Script para extraer francos de la grilla GIGANET y generar un Excel compatible con la app de calendario.
Busca celdas con "F" o "FRANCO" y mapea los días basándose en las fechas de los encabezados.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import json
import re

# Ruta del archivo original
ARCHIVO_ORIGINAL = '/workspaces/CalendarioApp/app-francos/src/app/3.Nueva Grilla 2026 GIGANET.xlsx'
ARCHIVO_SALIDA = '/workspaces/CalendarioApp/app-francos/src/assets/francos.json'

def extraer_fecha_de_encabezado(texto):
    """
    Extrae una fecha del texto del encabezado en formato YYYY-MM-DD.
    """
    if not texto:
        return None
    
    texto = str(texto).strip()
    
    # Patrón: YYYY-MM-DD (podría tener timestamp al final)
    match = re.search(r'(\d{4})-(\d{2})-(\d{2})', texto)
    if match:
        fecha_str = match.group(0)  # "2026-05-25"
        return fecha_str
    
    return None

def obtener_nombre_empleado(ws, row_idx):
    """Obtiene el nombre del empleado de la columna B de una fila."""
    nombre = ws.cell(row_idx, 2).value
    if nombre:
        return str(nombre).strip()
    return None

def procesar_hoja(wb, nombre_hoja):
    """
    Procesa una hoja individual para extraer empleados y sus francos.
    
    ESTRUCTURA:
    - Columna B: Nombre del empleado
    - Fila 3: Fechas de los días
    - Datos: "P" = Turno presente, NÚMERO o VACÍO = FRANCO
    """
    print(f"\n📋 Procesando hoja: '{nombre_hoja}'")
    
    try:
        ws = wb[nombre_hoja]
    except KeyError:
        print(f"  ❌ Hoja no encontrada")
        return {}
    
    empleados_francos = {}
    
    # Paso 1: Extraer fechas de la fila 3
    fechas_columnas = {}
    for col_idx in range(3, ws.max_column + 1):  # Comienza en columna C
        fecha_str = extraer_fecha_de_encabezado(ws.cell(3, col_idx).value)
        if fecha_str:
            if fecha_str not in fechas_columnas:
                fechas_columnas[fecha_str] = []
            fechas_columnas[fecha_str].append(col_idx)
    
    if not fechas_columnas:
        print(f"  ⚠️  No se encontraron fechas en la fila 3")
        return {}
    
    print(f"  ✓ Se encontraron {len(fechas_columnas)} fechas únicas")
    
    # Paso 2: Procesar empleados (a partir de fila 7)
    for row_idx in range(7, ws.max_row + 1):
        nombre = obtener_nombre_empleado(ws, row_idx)
        
        if not nombre:
            continue
        
        # Limpiar nombre (remove números iniciales como "1 GIANLUCA" -> "GIANLUCA")
        nombre = nombre.strip()
        nombre_limpio = nombre
        
        # Remover números iniciales
        if nombre and nombre[0].isdigit():
            for i, char in enumerate(nombre):
                if not char.isdigit() and char != ' ':
                    nombre_limpio = nombre[i:].strip()
                    break
        
        if not nombre_limpio or nombre_limpio.upper() in ['*', '']:
            continue
        
        francos = []
        
        # Para cada fecha
        for fecha_key in sorted(fechas_columnas.keys()):
            columnas_del_dia = fechas_columnas[fecha_key]
            
            # Verificar si tiene "P" (turno presente) en alguna columna del día
            tiene_presente = False
            
            for col_idx in columnas_del_dia:
                celda_valor = ws.cell(row_idx, col_idx).value
                
                # Solo "P" indica que tiene turno asignado
                if celda_valor and str(celda_valor).strip().upper() == 'P':
                    tiene_presente = True
                    break
            
            # Si NO tiene "P", es franco (aunque tenga número u otro marcador)
            if not tiene_presente:
                # Convertir fecha string a datetime para mantener consistencia
                try:
                    fecha = datetime.strptime(fecha_key, '%Y-%m-%d')
                    francos.append(fecha)
                except ValueError:
                    pass
        
        if francos:
            if nombre_limpio not in empleados_francos:
                empleados_francos[nombre_limpio] = []
            empleados_francos[nombre_limpio].extend(francos)
    
    print(f"  ✓ Se extrajeron {len(empleados_francos)} empleados con francos")
    
    return empleados_francos

def crear_archivo_compatible(todos_empleados_francos):
    """
    Crea un archivo JSON compatible con la app de Angular.
    """
    
    print("\n📝 Creando JSON compatible...")
    
    datos_json = []
    for nombre, francos in sorted(todos_empleados_francos.items()):
        fila = {"Nombre": nombre}
        for fecha in francos:
            fila[fecha.strftime('%Y-%m-%d')] = "F"
        datos_json.append(fila)
        
    try:
        with open(ARCHIVO_SALIDA, 'w', encoding='utf-8') as f:
            json.dump(datos_json, f, ensure_ascii=False, indent=2)
        print(f"\n✅ JSON generado exitosamente: {ARCHIVO_SALIDA}")
        return True
    except Exception as e:
        print(f"\n❌ Error al guardar: {e}")
        return False

def main():
    """Función principal."""
    print("="*80)
    print("🔄 EXTRAYENDO FRANCOS DEL EXCEL ORIGINAL")
    print("="*80)
    
    try:
        wb_original = openpyxl.load_workbook(ARCHIVO_ORIGINAL)
    except FileNotFoundError:
        print(f"❌ Archivo no encontrado: {ARCHIVO_ORIGINAL}")
        return
    except Exception as e:
        print(f"❌ Error al abrir el archivo: {e}")
        return
    
    print(f"\n✓ Archivo abierto correctamente")
    print(f"✓ Hojas disponibles: {len(wb_original.sheetnames)}")
    
    # Hojas relevantes (las que contienen fechas y turnos)
    hojas_relevantes = [h for h in wb_original.sheetnames if ' AL ' in h]
    
    print(f"✓ Hojas con fechas encontradas: {len(hojas_relevantes)}")
    
    # Procesar cada hoja
    todos_empleados_francos = {}
    
    for nombre_hoja in hojas_relevantes:
        empleados_francos = procesar_hoja(wb_original, nombre_hoja)
        
        # Combinar resultados
        for nombre, francos in empleados_francos.items():
            if nombre not in todos_empleados_francos:
                todos_empleados_francos[nombre] = []
            todos_empleados_francos[nombre].extend(francos)
    
    # Crear archivo compatible
    if todos_empleados_francos:
        crear_archivo_compatible(todos_empleados_francos)
        print("\n" + "="*80)
        print("✅ PROCESO COMPLETADO")
        print("="*80)
    else:
        print("\n❌ No se encontraron francos en ninguna hoja")

if __name__ == "__main__":
    main()
